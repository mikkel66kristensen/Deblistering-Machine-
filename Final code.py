# Importér pakker 
from enum import Enum   #pakke til at definere states
import cv2              #computer vision
import numpy as np      #databehandling
import ids_peak.ids_peak as ids_peak                          #kamera
import ids_peak.ids_peak_ipl_extension as ids_ipl_extension   #kamera
from openpyxl import load_workbook  #excel pakke

# Definer states
class State(Enum):
    INIT = "init"
    IMAGE = "image"
    CIRKEL = "cirkel"
    REKTANGEL = "rektangel"
    MOTOR = "motor"
    CALIBRATE = "calibrate"
    DATABASE = "database"
    DELETEPRESET = "deletepreset"


# Lav klasse, start state og globale konstanter
class StateMachine:
    def __init__(self):
        self.state = State.INIT

        # "globale" værdier (None - starter tomme)
        self.avg_x_spacing = None
        self.avg_y_spacing = None
        self.spacing = None
        self.maxspacing = None
        self.avg_radius = None
        self.avg_width = None    
        self.avg_height = None
        self.max_dist = None              
        self.mm_pr_pixel = None
        self.img = cv2.imread(r"C:\Users\mikke\Desktop\Projekt\rund4.png")
        self.imgrekt = cv2.imread(r"C:\Users\mikke\Desktop\Projekt\aflang5.png")
        self.imgraw = None
        self.sn_kode = None
        self.wb = load_workbook(r"C:\Users\mikke\Desktop\Kode\Pille database.xlsx")
        self.ws = self.wb["Sheet1"]
        self.sn_kode = None
    
# Definér hvordan states ændres i praksis
    def transition_to(self, new_state):
        # 1. exit gammel state
        exit_method = f"on_exit_{self.state.value}"   #funktion hedder on_exit_{state}
        if hasattr(self, exit_method):                #hvis self indeholder exit_metod -> true
            getattr(self, exit_method)()              #henter funktionen exit_metod og kalden funktion ved () 

        # 2. skift state
        self.state = new_state
        print(f"State: {self.state.value}")

        # 3. enter ny state
        enter_method = f"on_enter_{self.state.value}"
        if hasattr(self, enter_method):
            getattr(self, enter_method)()

# Definér hvordan der skrives i excel
    def write_in_excel(self):  
        sn_kode = self.sn_kode
        wb = load_workbook(r"C:\Users\mikke\Desktop\Kode\Pille database.xlsx")
        ws = wb["Sheet1"]
        
        spacing = self.spacing
        maxspacing = self.maxspacing
        radius = self.avg_radius
        width = self.avg_width   
        height = self.avg_height
           
        
        fundet = False
        
        if sn_kode is not None:
            for row in ws.iter_rows(min_col=1, max_col=1):
                cell = row[0]

                if not cell.value:
                    continue

                if str(cell.value).strip() == sn_kode:
                    print(f"Fundet i række {cell.row}")
                    fundet = True
                    row = cell.row
                    break

        if not fundet:
            print("Ikke fundet")
            col = 1

            for row in range(1, ws.max_row + 2):  # +2 så vi også kan skrive i ny række
                cell = ws.cell(row=row, column=col)

                if cell.value is None:
                    # celle
                    cell.value = sn_kode
                    
                    # cellen ved siden af (kolonne + 1)
                    cell_nabo1 = ws.cell(row=row, column=col + 1)
                    cell_nabo1.value = spacing
                    
                    # cellen 2 ved siden af (kolonne + 2)
                    cell_nabo2 = ws.cell(row=row, column=col + 2)
                    cell_nabo2.value = maxspacing   
                    
                    # cellen 3 ved siden af (kolonne + 3)
                    cell_nabo3 = ws.cell(row=row, column=col + 3)
                    cell_nabo3.value = radius

                    # cellen 4 ved siden af (kolonne + 4)
                    cell_nabo4 = ws.cell(row=row, column=col + 4)
                    cell_nabo4.value = width
 
                    # cellen 5 ved siden af (kolonne + 5)
                    cell_nabo5 = ws.cell(row=row, column=col + 5)
                    cell_nabo5.value = height
                    
            
                    
        
                    # Save værdier i excel
                    wb.save(r"C:\Users\mikke\Desktop\Kode\Pille database.xlsx")
                    print("Værdier er gemt i database")

        if fundet and self.avg_x_spacing is not None and self.avg_y_spacing is not None:
            col = 1
            print("Fundet") 
            key = input("Vil du overskrive eksisterende værdier: y eller n")
                       
            if key == "y":
                # cellen ved siden af (kolonne + 1)
                cell_nabo1 = ws.cell(row=row, column=col + 1)
                cell_nabo1.value = self.spacing
                    
                # cellen 2 ved siden af (kolonne + 2)
                cell_nabo2 = ws.cell(row=row, column=col + 2)
                cell_nabo2.value = self.maxspacing 
                
                # cellen 3 ved siden af (kolonne + 3)
                cell_nabo3 = ws.cell(row=row, column=col + 3)
                cell_nabo3.value = self.avg_radius                 
                
                # cellen 4 ved siden af (kolonne + 4)
                cell_nabo4 = ws.cell(row=row, column=col + 4)
                cell_nabo4.value = self.avg_width

                # cellen 5 ved siden af (kolonne + 5)
                cell_nabo5 = ws.cell(row=row, column=col + 5)
                cell_nabo5.value = self.avg_height
                
         
                 


                
                # Save værdier i excel
                wb.save(r"C:\Users\mikke\Desktop\Kode\Pille database.xlsx")
                print("Værdier er gemt i database")    
                
            elif key == "n":
                print("Værdier overskrives ikke")
            else:
                print("Invalid input - værdier overskrives ikke")
                    
                    


# Definer hvad states gør
  
    def on_enter_init(self):            #udfyldt
        self.avg_x_spacing = None
        self.avg_y_spacing = None
        self.spacing = None
        self.maxspacing = None
        self.mm_pr_pixel = None
        self.avg_height = None
        self.avg_width = None        
        self.avg_radius = None
        self.max_dist = None
        self.sn_kode = None 
        self.imgraw = None


    def on_enter_image(self):           #udfyldt 
        ids_peak.Library.Initialize()
        device_manager = ids_peak.DeviceManager.Instance()
        device_manager.Update()
        device_descriptors = device_manager.Devices()

        print("Found Devices: " + str(len(device_descriptors)))
        for device_descriptor in device_descriptors:
            print(device_descriptor.DisplayName())
            
        # Open camera
        device = device_descriptors[0].OpenDevice(ids_peak.DeviceAccessType_Control)
        print("Opened Device: " + device.DisplayName())
        remote_device_nodemap = device.RemoteDevice().NodeMaps()[0]
            
        # Trigger mode
        remote_device_nodemap.FindNode("TriggerSelector").SetCurrentEntry("ExposureStart")
        remote_device_nodemap.FindNode("TriggerSource").SetCurrentEntry("Software")
        remote_device_nodemap.FindNode("TriggerMode").SetCurrentEntry("On")
        
                
        # Start image accusition
        datastream = device.DataStreams()[0].OpenDataStream()
        payload_size = remote_device_nodemap.FindNode("PayloadSize").Value()
        for i in range(datastream.NumBuffersAnnouncedMinRequired()):
            buffer = datastream.AllocAndAnnounceBuffer(payload_size)
            datastream.QueueBuffer(buffer)
            
        datastream.StartAcquisition()
        remote_device_nodemap.FindNode("AcquisitionStart").Execute()
        remote_device_nodemap.FindNode("AcquisitionStart").WaitUntilDone()
        
        

        # Trigger image
        remote_device_nodemap.FindNode("TriggerSoftware").Execute()
        buffer = datastream.WaitForFinishedBuffer(1000)
        
        ipl_image = ids_ipl_extension.BufferToImage(buffer)

        self.imgraw = ipl_image.get_numpy().copy() 


        print(self.imgraw.dtype, self.imgraw.min(), self.imgraw.max())
        print(np.percentile(self.imgraw, [0, 1, 50, 99, 100])) # printer saturation 0 til 255
        print(self.imgraw.shape)
        
        # Vis billede 
        resize_raw = cv2.resize(self.imgraw,(800,600))
        cv2.imshow("RAW billede", resize_raw)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        

        # Terminate
        try:
            if datastream is not None:
                try:
                    remote_device_nodemap.FindNode("AcquisitionStop").Execute()
                except:
                    pass

                datastream.StopAcquisition()
                datastream.Flush(ids_peak.DataStreamFlushMode_DiscardAll)
                datastream.RevokeAllBuffers()
        except:
            pass
        
        
        ids_peak.Library.Close()
        
        
        
        
        
        
        
        
        
        
        
        
        
        
  
        
  
        
    def on_enter_cirkel(self):          #udfyldt
      
       # Tjek om billede findes 
        if self.imgraw is None:
            print("Intet billede indlæst")
            return
        
        
        # Læs billede  
        img = self.imgraw    
 #      gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        gray = img
        resize_gray = cv2.resize(gray,(800,600))
        blur_gray = cv2.GaussianBlur(resize_gray, (9,9), 2)
        
        # Vis gråbillede
        cv2.imshow("blur billede", blur_gray)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        
        # Find cirkler
        circles = cv2.HoughCircles(
        blur_gray,
        cv2.HOUGH_GRADIENT,
        dp=1.8,             #opløsning af billede, 1 er samme som billede, høj er højere 
        minDist=25,
        param1=105,          #edge detection - lav værdi mere støj (50-150)
        param2=75,          #tærskel for at acceptere cirkel - lav værdi flere cirkler (20-100) 
        minRadius=12,
        maxRadius=75
        )
        
        if circles is None:
            print ("Ingen cirkler fundet")
            return
            
        else:
            # Fjern cirkler som ligger uden for størrelse grænse
            radii = circles[0, :, 2]         #udtræk radius for hver cirkel
            median_r = np.median(radii)      #beregn median
            lower = 0.75 * median_r           #beregn nedre grænse 
            upper = 1.25 * median_r           #beregn øvre grænse
     
            filtered_circles = [
                c for c in circles[0, :]
                if lower <= c[2] <= upper]      #filtrer cirkeler fra 
        
        
            # Definér cirkel radius
            mean_r = np.mean([c[2] for c in filtered_circles])                    #gennemsnit af radii
            normalized_circles = [(c[0], c[1], mean_r) for c in filtered_circles]  #definer radius


            # Definér cirkel position
            tolerance = 10
        
                # x-justering
            circles_sorted_x = sorted(normalized_circles, key=lambda c: c[0])
            x_groups = []
            current_group = [circles_sorted_x[0]]
            
            for c in circles_sorted_x[1:]:
                if abs(c[0] - current_group[-1][0]) < tolerance:
                        current_group.append(c)
                else:
                    x_groups.append(current_group)
                    current_group = [c]
                    x_groups.append(current_group)

            adjusted = []
            column_centers = []

            for group in x_groups:
                avg_x = int(np.mean([c[0] for c in group]))
                column_centers.append(avg_x)
                for (x, y, r) in group:
                    adjusted.append([avg_x, y, r])
            
            
            
                # y-justering
                adjusted_sorted_y = sorted(adjusted, key=lambda c: c[1])
                y_groups = []
                current_group = [adjusted_sorted_y[0]]

            for c in adjusted_sorted_y[1:]:
                if abs(c[1] - current_group[-1][1]) < tolerance:
                        current_group.append(c)
                else:
                    y_groups.append(current_group)
                    current_group = [c]
                    y_groups.append(current_group)

            final_positions = []
            row_centers = []

            for group in y_groups:
                avg_y = int(np.mean([c[1] for c in group]))
                row_centers.append(avg_y)
                for (x, y, r) in group:
                    final_positions.append([x, avg_y, r])
        
            # Definer vektorer med positioner er centrum
            column_centers = sorted(column_centers)
            row_centers = sorted(row_centers)
        
            
            # Fjern dubletter i vektorer
            column_centers = sorted(set(column_centers))    
            row_centers = sorted(set(row_centers))    
        
            print(column_centers)
            print(row_centers)
            
            # Beregn spacings
            diffs_x = np.abs(np.diff(column_centers))
            diffs_y = np.abs(np.diff(row_centers))
            
    #        print(diffs_x)
    #       print(diffs_y)  
    #        print("mean radius", mean_r)
            
            #Beregn de brugbare spacings
            if len(column_centers) > len(row_centers) and min(diffs_x) > 2*mean_r:
                spacing = np.mean(diffs_x)
                maxspacing = max(column_centers)-min(column_centers)
                print("større en 2*radius")
            else:
                spacing = np.mean(diffs_y) 
                maxspacing = max(row_centers)-min(row_centers)
        
        
            if self.mm_pr_pixel is not None:    
                self.spacing = spacing * self.mm_pr_pixel
                self.maxspacing = maxspacing * self.mm_pr_pixel
                
    #            print(self.spacing)
    #             print(self.maxspacing)
            
            
            
            # tæl cirkler
            count = len(filtered_circles)
        
            # tegn cirkler på billedet
           # img = resize_gray.copy()                   # laver kopi af billede, så råbillide forbliver
            img = resize_gray.copy()
            if len(img.shape) == 2:
                    img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
        
            for (x, y, r) in final_positions:
                cv2.circle(img, (int(x), int(y)), int(r), (0, 0, 255), 2)  
                cv2.circle(img, (int(x), int(y)), 2, (255, 0, 0), 3)      
                 
            if self.mm_pr_pixel is not None:
                print(f"Antal cirkler: {count}")
                print(f"Spacing: {self.spacing} mm")
                print(f"Max spacing: {self.maxspacing} mm")
            else:
                print("kalibrer for at beregne spacing")
        
            
         
            
                 # Write in excel
            if self.mm_pr_pixel is not None and self.sn_kode is not None:
                    self.avg_radius = mean_r * self.mm_pr_pixel
                    self.avg_width = None
                    self.avg_height = None
                    self.spacing = spacing * self.mm_pr_pixel
                    self.maxspacing = maxspacing * self.mm_pr_pixel   
                    self.write_in_excel()

        
       
            # Vis figur
            resize_img = cv2.resize(img,(800,600))
            cv2.imshow("Circles with Spacing", resize_img)
            cv2.waitKey(0)
            cv2.destroyAllWindows()
        
        

            
        
        
        
        
    
    def on_enter_rektangel(self):       #udfyldt
       # Tjek om billede findes 
        if self.imgraw is None:
            print("Intet billede indlæst")
            return
    
        img = self.imgraw
 #      gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        gray = img
        resize_gray = cv2.resize(gray,(800,600))
        blur_gray = cv2.GaussianBlur(resize_gray, (9,9), 2)
        
        cv2.imshow("blur billede", blur_gray)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        
        # Læs billede 

        
        
        
        # Find kanter
## threshhold (0=sort, 255=hvid) (_, giver kun 2. værdi)
        thresh = cv2.adaptiveThreshold(blur_gray, 255, 
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C, 
        cv2.THRESH_BINARY_INV, 
        blockSize=5,  # størrelse af lokalt område (skal være ulige)
        C=2)           # konstant der trækkes fra


        
        
      
        cv2.imshow("thresh billede", thresh)
        cv2.waitKey(0)
        cv2.destroyAllWindows()

        
        
        # Find konturer
        contours, _ = cv2.findContours(thresh, cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)

        rectangles = []                         # laver tom liste
        areas = []                              # laver tom liste
        
        # Find rektangler, frasortér små, og justér vinkel
        for cnt in contours:
            rect = cv2.minAreaRect(cnt)
            (cx, cy), (w, h), angle = rect     # (centrum),(bredde-højde),vinkel    
            
            # Fjern små
            sizelimitmin = 10
            sizelimitmax = 99999
            
           
            img_area = 800 * 600
            if w * h > img_area * 0.2:  # Spring over hvis kontur fylder >90% af billedet
                continue    
            if w < sizelimitmin or h < sizelimitmin:
                continue
            if w > sizelimitmax or h > sizelimitmax:
                continue
            
            # Gør bredden til den længste side
            if w < h:
                w, h = h, w
                angle += 90
                
            # Justér vinkel
            if angle < -90:
                angle += 180
            elif angle > 90:
                angle -= 180
                
            # Håndtér data    
            area = w * h                  #Beregner areal
            areas.append(area)            #Opbevarer værdi i listen areas
            rectangles.append([int(cx), int(cy), w, h, angle, area])  #Gemmer information
            
            
            
        #Filtrér rektangler
        
        if areas:
            median_area = np.median(areas)
            min_area = 0.5 * median_area
            max_area = 1.5 * median_area
            rectangles = [r for r in rectangles if min_area <= r[5] <= max_area]  # beholde kun rekt indenfor interval
        
        if not rectangles:
            print("Ingen rektangler fundet")
            return
        
        #Bestem dominant vinkel 
        zero_count = sum(1 for r in rectangles if -25 <= r[4] <= 25)  # tæller 0deg rekt
        ninety_count = sum(1 for r in rectangles if 65 <= abs(r[4]) <= 115) #tæller 90deg rekt
        dominant_angle = 0 if zero_count >= ninety_count else 90 # vurderer om 0 eller 90 er dominant

        #Normalisér vinkler som er tæt på 0 eller 90deg
        non_normalized_angles = []
        for i, r in enumerate(rectangles):
            a = r[4]
            if dominant_angle == 0 and -25 <= a <= 25:
                rectangles[i][4] = 0
            elif dominant_angle == 90 and 65 <= abs(a) <= 115:
                rectangles[i][4] = 90
            else:
                non_normalized_angles.append(a)  # gem de vinkler, der ikke blev 0 eller 90
       
        #Beregn vinkel for ikke normaliserede vinkler
        if non_normalized_angles:
            avg_angle = np.mean(non_normalized_angles)
            # Sæt disse vinkler til gennemsnittet
            for i, r in enumerate(rectangles):
                a = r[4]
                if a not in [0, 90]:
                    rectangles[i][4] = avg_angle
        
        #Find gennemsnits bredde og højde
        avg_w = np.mean([r[2] for r in rectangles])
        avg_h = np.mean([r[3] for r in rectangles])
        
        #Grid justér
        tolerance = 20 
        
            # Justér x
        rects_sorted_x = sorted(rectangles, key=lambda r: r[0])
        x_groups = []
        current_group = [rects_sorted_x[0]]
        for r in rects_sorted_x[1:]:
            if abs(r[0] - current_group[-1][0]) < tolerance:
                current_group.append(r)
            else:
                x_groups.append(current_group)
                current_group = [r]
        x_groups.append(current_group)

        adjusted = []
        column_centers = []
        for group in x_groups:
            avg_x = int(np.mean([r[0] for r in group]))
            column_centers.append(avg_x)
            for (x, y, w, h, a, area) in group:
                adjusted.append([avg_x, y, w, h, a])

            # Justér y
        adjusted_sorted_y = sorted(adjusted, key=lambda r: r[1])
        y_groups = []
        current_group = [adjusted_sorted_y[0]]
        for r in adjusted_sorted_y[1:]:
            if abs(r[1] - current_group[-1][1]) < tolerance:
                current_group.append(r)
            else:
                y_groups.append(current_group)
                current_group = [r]
        y_groups.append(current_group)

        final_rects = []
        row_centers = []
        for group in y_groups:
            avg_y = int(np.mean([r[1] for r in group]))
            row_centers.append(avg_y)
            for (x, y, w, h, a) in group:
                final_rects.append((x, avg_y, avg_w, avg_h, a))
                
        # Beregn spacing 
        avg_x_spacing = int(np.mean(np.diff(sorted(column_centers)))) if len(column_centers) > 1 else 0
        avg_y_spacing = int(np.mean(np.diff(sorted(row_centers)))) if len(row_centers) > 1 else 0
        
        
        # Beregn maxspacing i x- og y-retning
        max_x_spacing = max(column_centers)-min(column_centers)
        max_y_spacing = max(row_centers)-min(row_centers)
        max_spacing = max(max_x_spacing,max_y_spacing)

        
   
           #Definér globalt
        if self.mm_pr_pixel is not None:    
            self.avg_x_spacing = avg_x_spacing * self.mm_pr_pixel            #definerer globalt
            self.avg_y_spacing = avg_y_spacing * self.mm_pr_pixel 
        
            
       
        # Print
        if self.mm_pr_pixel is not None:
            print(f"Antal rektangler: {len(rectangles)}")
            print(f"Gennemsnitlig X-spacing: {self.avg_x_spacing} mm")
            print(f"Gennemsnitlig Y-spacing: {self.avg_y_spacing} mm")
            
        
            # Write in excel
            self.avg_radius = None
            self.avg_width = avg_w * self.mm_pr_pixel
            self.avg_height = avg_h * self.mm_pr_pixel
            self.max_dist = max_spacing * self.mm_pr_pixel
            if self.sn_kode is not None:
                self.write_in_excel()
        else:
                print("Kalibrer for at beregne spacing")
       
        
       # Tegn figur      
        img = cv2.cvtColor(resize_gray, cv2.COLOR_GRAY2BGR)
        for (x, y, w, h, a) in final_rects:
            box = cv2.boxPoints(((x, y), (w, h), a))
            box = np.int32(box)
            cv2.drawContours(img, [box], 0, (0, 255, 0), 2)
            cv2.circle(img, (int(x), int(y)), 2, (0, 0, 255), 3)
        
        cv2.imshow("Rektangel billede", img)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
       
        

        
       
        
    def on_enter_calibrate(self):       #udfyldt
         
       # Tjek om billede findes 
        if self.imgraw is None:
            print("Intet billede indlæst")
            return       
 
        
 # Load billede
        img = self.imgraw
 #      gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        gray = img
        resize_gray = cv2.resize(gray,(800,600))
        blur_gray = cv2.GaussianBlur(resize_gray, (5,5), 2)
        
        cv2.imshow("blur billede", blur_gray)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
       
        

        
        

        # Bed bruger om kalibreringsdiameter i mm
        while True:
            try:
                calib_diameter_mm = float(input("Indtast kalibrerings-cirkeldiameter i mm: "))
                break  # stopper loopet når input er gyldigt
            except ValueError:
                print("Ugyldigt input, brug tal.")
        

        # Find cirkler
        circles = cv2.HoughCircles(
        blur_gray,
        cv2.HOUGH_GRADIENT,
        dp=1.2,             #opløsning af billede, 1 er samme som billede, høj er højere 
        minDist=200,
        param1=80,          #edge detection - lav værdi mere støj (50-150)
        param2=60,          #tærskel for at acceptere cirkel - lav værdi flere cirkler (20-100) 
        minRadius=150,
        maxRadius=400
        )
        
    
    
        # Formatér cirkler
        if circles is not None:
            circles = circles[0]
            for (x, y, r) in circles:
                    x, y, r = int(x), int(y), int(r)

            # Kalibrering
            radius_px = int(circles[0][2])
            mm_per_pixel = calib_diameter_mm / (2 * radius_px)
            print("mm/pixel: ", mm_per_pixel)

            # Definér globalt
            self.mm_pr_pixel = mm_per_pixel

            # Tegn cirklen
            img = cv2.cvtColor(resize_gray, cv2.COLOR_GRAY2BGR)
            cv2.circle(img, (int(x), int(y)), int(r), (0, 0, 255), 2)  
            cv2.circle(img, (int(x), int(y)), 2, (255, 0, 0), 3) 

            # Skriv tekst på billede
            cv2.putText(img, f"mm/pixel: {mm_per_pixel:.4f}", (10, 30),
                        cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 0), 2)  # sort baggrund
        
            # Vis billede
           
            cv2.imshow("Circles Calibration", img)
            cv2.waitKey(0)
            cv2.destroyAllWindows()
            
        else: 
            print("Ingen cirkel fundet")
        
        
    def on_enter_database(self):
        print("database")
        sn_kode = input("Scan sn-kode").strip()



        fundet = False



        for row in self.ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]

            if cell.value is None:
                continue

            if str(cell.value).strip() == sn_kode:
                print(f"Fundet i række {cell.row}")
                fundet = True
                fundet_row = cell.row 

        if not fundet:
            print("Ikke fundet")
            
         
        if fundet:
            self.avg_x_spacing = self.ws.cell(row=fundet_row, column=2).value
            self.avg_y_spacing = self.ws.cell(row=fundet_row, column=3).value
            self.avg_radius = self.ws.cell(row=fundet_row, column=4).value
            self.avg_width = self.ws.cell(row=fundet_row, column=5).value
            self.avg_height = self.ws.cell(row=fundet_row, column=6).value
            self.max_dist = self.ws.cell(row=fundet_row, column=7).value
            
            
            input("Preset found - Tryk Enter for at gå videre")
            sm.transition_to(State.MOTOR)
            
        if not fundet:
            if self.mm_pr_pixel is not None:
                input("Preset not found - Tryk Enter for at tage billede")
                sm.transition_to(State.IMAGE)
            else:
                input("Preset not found - Tryk Enter for at kalibrere kamera")
                sm.transition_to(State.CALIBRATE)
            
        self.sn_kode = sn_kode
        
    def on_enter_motor(self):
        print("MOTOR")
        
        
        
    def on_enter_deletepreset(self):    #udfyldt
        wb = load_workbook(r"C:\Users\mikke\Desktop\Kode\Pille database.xlsx")
        ws = wb["Sheet1"]
        
        print("delete preset")
        sn_kode = input("Scan sn-kode").strip()
        
        col = 1

        fundet = False

        for row in self.ws.iter_rows(min_col=1, max_col=1):
            cell = row[0]

            if cell.value is None:
                continue

            if str(cell.value).strip() == sn_kode:
                print(f"Fundet i række {cell.row}")
                fundet = True
                found_row = cell.row 

        if not fundet:
            print("Ikke fundet")
            
         
        if fundet:
            cell = ws.cell(row=found_row, column=col)            
            cell.value = None
            
            # cellen ved siden af (kolonne + 1)
            cell_nabo1 = ws.cell(row=found_row, column=col + 1)
            cell_nabo1.value = None
            
            # cellen 2 ved siden af (kolonne + 2)
            cell_nabo2 = ws.cell(row=found_row, column=col + 2)
            cell_nabo2.value = None
            
            # cellen 3 ved siden af (kolonne + 3)
            cell_nabo3 = ws.cell(row=found_row, column=col + 3)
            cell_nabo3.value = None
            
            # cellen 4 ved siden af (kolonne + 4)
            cell_nabo4 = ws.cell(row=found_row, column=col + 4)
            cell_nabo4.value = None
            
            # cellen 5 ved siden af (kolonne + 5)
            cell_nabo5 = ws.cell(row=found_row, column=col + 5)
            cell_nabo5.value = None
            
            # cellen 6 ved siden af (kolonne + 6)
            cell_nabo6 = ws.cell(row=found_row, column=col + 6)
            cell_nabo6.value = None
            
            # Save værdier i excel
            wb.save(r"C:\Users\mikke\Desktop\Kode\Pille database.xlsx")
            print("Værdier er slettet i database")
        
        
        
        

#Styring "main loop"
sm = StateMachine()

while True:
    key = input("instruks her ")
 
    if key == "0":
        sm.transition_to(State.INIT)
    elif key == "1":
        sm.transition_to(State.IMAGE)
    elif key == "2":
        sm.transition_to(State.CIRKEL)
    elif key == "3":
        sm.transition_to(State.REKTANGEL)
    elif key == "4":
        sm.transition_to(State.MOTOR)            
    elif key == "c":
        sm.transition_to(State.CALIBRATE)
    elif key == "d":
        sm.transition_to(State.DATABASE)                
    elif key == "p":                               #printer state 
        print("state:",sm.state.value)
        print("avg x spacing:",sm.avg_x_spacing)
        print("avg y spacing:",sm.avg_y_spacing) 
        print("avg radius:",sm.avg_radius)
        print("avg width:",sm.avg_width)        
        print("avg height:",sm.avg_height)
        print("max dist:",sm.max_dist)          
        print("mm pr. pixel:",sm.mm_pr_pixel) 
        print("SN-kode:",sm.sn_kode) 
    elif key== "b":                                #printer rå billede
        cv2.imshow("RAW", sm.imgraw)
        cv2.waitKey(0)
        cv2.destroyAllWindows()
        
       
         
      
        
    elif key == "9":
        sm.transition_to(State.DELETEPRESET) 
        
    elif key == "q":
            break